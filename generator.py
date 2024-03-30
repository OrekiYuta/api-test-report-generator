import glob
import os
import shutil
import traceback
import warnings

import pandas as pd
from math import ceil
from PIL import Image, ImageFont, ImageDraw
from openpyxl import load_workbook, drawing
from utils.PathManager import load_path_manager as lpm

PIL_IMAGE_MODE = 'RGB'
PIL_WIDTH_INDEX = 0
PIL_HEIGHT_INDEX = 1
ROLE_SECURITY_POSITION = {
    "AA": (7, 4),
    "BB": (4, 5),
    "MEMBER": (3, 6),
    "ER": (5, 7),
    "GUEST": (8, 8),
    "ADMIN": (6, 9),
    "SYSTEM": (9, 10)
}

file_master_data = lpm.input("API Test Master Data.xlsx")
file_output_template = lpm.template("API Test Template.xlsx")
folder_input_api_result = str(lpm.input("API-Result"))
folder_input_sql_statement = str(lpm.input("SQL-Statement"))
folder_input_jaeger = str(lpm.input("Jaeger-Screenshot"))
folder_input_junit = str(lpm.input("Junit-Screenshot"))
folder_input_db_schema = str(lpm.input("DB-Schema"))
folder_output = str(lpm.output)
folder_input = str(lpm.input)


def break_fix(text, width, font, draw):
    if not text:
        return
    lo = 0
    hi = len(text)
    while lo < hi:
        mid = (lo + hi + 1) // 2
        t = text[:mid]
        w, h = draw.textsize(t, font=font)
        if w <= width:
            lo = mid
        else:
            hi = mid - 1
    t = text[:lo]
    w, h = draw.textsize(t, font=font)
    yield t, w, h
    yield from break_fix(text[lo:], width, font, draw)


def textfile_to_image(textfile_path, max_width=None):
    try:
        """Convert text file to a grayscale image.

        arguments:
        textfile_path - the content of this file will be converted to an image
        font_path - path to a font file (for example impact.ttf)
        """
        # parse the file into lines stripped of whitespace on the right side
        with open(textfile_path) as text_file:
            lines = tuple(line.rstrip() for line in text_file.readlines())

        # choose a font (you can see more detail in the linked library on github)
        font = None
        large_font = 20  # get better resolution with larger size
        font = ImageFont.load_default()  # Note pillow version == 9.5.0

        # make a sufficiently sized background image based on the combination of font and lines
        font_points_to_pixels = lambda pt: round(pt * 96.0 / 72)
        margin_pixels = 20

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")

            # height of the background image
            tallest_line = max(lines, key=lambda line: font.getsize(line)[PIL_HEIGHT_INDEX])
            max_line_height = font_points_to_pixels(font.getsize(tallest_line)[PIL_HEIGHT_INDEX])
            realistic_line_height = max_line_height * 0.8  # apparently it measures a lot of space above visible content

            image_height = int(ceil(realistic_line_height * len(lines) + 2 * margin_pixels))

            # width of the background image
            widest_line = max(lines, key=lambda s: font.getsize(s)[PIL_WIDTH_INDEX])

            if max_width is None:
                max_line_width = font_points_to_pixels(font.getsize(widest_line)[PIL_WIDTH_INDEX])
            else:
                max_line_width = max_width

        image_width = int(ceil(max_line_width + (2 * margin_pixels)))

        # draw the background
        image = Image.new(PIL_IMAGE_MODE, (image_width, image_height), color="white")
        draw = ImageDraw.Draw(image)

        # draw each line of text
        font_color = 0  # black
        horizontal_position = margin_pixels
        for i, line in enumerate(lines):
            vertical_position = int(round(margin_pixels + (i * realistic_line_height)))
            draw.text((horizontal_position, vertical_position), line, fill=font_color, font=font)

        return image

    except Exception as e:
        # print(row["Scenario"])
        print(e)
        traceback.print_exc()


def add_image_to_excel(file_path, worksheet, img_anchor):
    excel_img = drawing.image.Image(file_path)
    excel_img.anchor = img_anchor
    worksheet.add_image(excel_img)


def generate_txt_file(sql_statement, api_id, scenario):
    file_path = f'{folder_input_sql_statement}/{api_id}/{scenario}.txt'

    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(sql_statement)


def convert_file_txt_to_image(source_folder, api_id, scenario):
    target_image_file = f'{source_folder}/{api_id}/{scenario}.png'
    origin_txt_file = f'{source_folder}/{api_id}/{scenario}.txt'

    image = textfile_to_image(origin_txt_file)
    image.save(target_image_file)

    return target_image_file


def start():
    sheets_dict = pd.read_excel(file_master_data, engine='openpyxl', sheet_name=None)

    # Note: Filter Execution
    # cc api  count: 2
    cvh_item = ['API-001', 'API-002']
    # bb api count: 25
    batch_item = ['BAT-001', 'BAT-002', 'BAT-003', 'BAT-017', 'BAT-024',
                  'BAT-025', 'BAT-004', 'BAT-005', 'BAT-008', 'BAT-007',
                  'BAT-009', 'BAT-018', 'BAT-010', 'BAT-019', 'BAT-006',
                  'BAT-012', 'BAT-021', 'BAT-013', 'BAT-022', 'BAT-011',
                  'BAT-020', 'BAT-014', 'BAT-015', 'BAT-023', 'BAT-016']

    # aa admin api count: 11
    admin_item = ['API-077', 'API-078', 'API-079', 'API-080',
                  'API-081', 'API-082', 'API-086', 'API-087',
                  'API-083', 'API-084', 'API-085']

    # dd common api  count: 62
    # ee-API-007 special handling
    common_item = ['API-007', 'API-021', 'API-022', 'API-071', 'API-023', 'API-024',
                   'API-026', 'API-072', 'API-027', 'API-028', 'API-029', 'API-063',
                   'API-019', 'API-020', 'API-069', 'API-006', 'API-068', 'API-008',
                   'API-009', 'API-016', 'API-017', 'API-018', 'API-044', 'API-045',
                   'API-074', 'API-046', 'API-047', 'API-048', 'API-041', 'API-042',
                   'API-043', 'API-049', 'API-050', 'API-051', 'API-075', 'API-058',
                   'API-059', 'API-060', 'API-076', 'API-038', 'API-039', 'API-040',
                   'API-052', 'API-053', 'API-054', 'API-055', 'API-056', 'API-057',
                   'API-061', 'API-062', 'API-010', 'API-011', 'API-013', 'API-014',
                   'API-031', 'API-032', 'API-033', 'API-034', 'API-035', 'API-036',
                   'API-003', 'API-004']

    # single check api
    # single_item = ['BAT-024']
    single_item = ['API-020']

    # all
    all_set = set(cvh_item + batch_item + admin_item + common_item)
    all_item = list(all_set)

    # change here
    process_item = all_item

    for name, sheet in sheets_dict.items():
        if sheet.empty or name == "README":
            continue

        if name == "MASTER":
            output_path_struct = {}
            sql_statement_map = {}
            for index, row in sheet.iterrows():
                output_path_struct[row['API ID']] = {
                    'API ID': row['API ID'],
                    'FS ID': row['FS ID'],
                    'Function Name': row['Function Name'],
                    'API Name': row['API Name']
                }

                api_id = row['API ID']
                sql_statement = row['SQL Statement']
                sql_statement_map[api_id] = sql_statement

        # Filter Execution
        if name not in process_item:
            continue

        if not os.path.exists(folder_output):
            os.makedirs(folder_output)

        # module categorize folder and file
        current_api = output_path_struct.get(name)
        folder_api_test = f'{folder_output}/' \
                          f'[{current_api["FS ID"]}] {current_api["Function Name"]}'

        if not os.path.exists(folder_api_test):
            os.makedirs(folder_api_test)

        output_file = f'{folder_api_test}/' \
                      f'[{current_api["API ID"]}] {current_api["API Name"]}.xlsx'

        # direct file
        # if not os.path.exists(folder_output):
        #     os.makedirs(folder_output)
        #
        # output_file = f"{folder_output}/{name}.xlsx"

        shutil.copy2(file_output_template, output_file)

        api_id = name
        print(f"//////////// process {name} //////////// ")
        wb = load_workbook(output_file)
        for index, row in sheet.iterrows():
            print(f'----> {row["Scenario"]}')

            # construct sql statement file
            sql_statement_content = sql_statement_map.get(row['API ID'])
            if sql_statement_content is None or str(sql_statement_content) == 'nan':
                print(" pls add sql statement in master data file")
                generate_txt_file("TODO", row['API ID'], row['Scenario'])
            else:
                generate_txt_file(sql_statement_content, row['API ID'], row['Scenario'])

            for ws in wb:
                if index == 1 and ws.title == "Security":
                    tick = ws.cell(13, 1).value

                    role = row["Role Required"].split("_")[-1]
                    row_index, column_index = ROLE_SECURITY_POSITION[role]
                    ws.cell(row_index, column_index).value = tick
                    ws.cell(row_index, 11).value = f"Only {role} user can access"

                    for ws_row in ws.iter_rows(1):
                        for cell in ws_row:
                            if cell.value == "<ENDPOINT_NO>":
                                ws.cell(row=cell.row, column=1).value = row["API Index"]
                    for ws_row in ws.iter_rows(2):
                        for cell in ws_row:
                            if cell.value == "<ENDPOINT_PATH>":
                                ws.cell(row=cell.row, column=2).value = row["API End Point"]

                    # clean tick
                    ws.cell(13, 1).value = None
                    ws.cell(14, 1).value = None

                elif ws.title == "Application Logic":
                    row_index = index + 2
                    ws.cell(row_index, 1).value = api_id
                    ws.cell(row_index, 2).value = row["API End Point"]
                    ws.cell(row_index, 3).value = row["Scenario"]
                    ws.cell(row_index, 4).value = row["Test Description"].strip()

                    description = str(row["Test Description"]) if row["Test Description"] else ""
                    logic = str(row["Test Case Logic"]) if row["Test Case Logic"] else ""

                    if row["Corresponding Fields"] is not None and str(row["Corresponding Fields"]) == "nan":
                        fields = ""
                    else:
                        fields = str(row["Corresponding Fields"])

                    ws.cell(row_index, 5).value = description + logic + fields
                    # ws.cell(row_index, 5).value = row["Test Description"] + row["Test Case Logic"] + row[
                    #     "Corresponding Fields"]

                    ws.cell(row_index, 6).value = "PASSED"

                elif ws.title == "Scenario Template":
                    scenario_ws = wb.copy_worksheet(ws)
                    scenario_ws.title = row["Scenario"]
                    scenario_ws.cell(1, 2).value = row["FS ID"]
                    scenario_ws.cell(2, 2).value = row["Function ID"]
                    scenario_ws.cell(3, 2).value = row["Function Name"]
                    scenario_ws.cell(5, 1).value = row["Test Description"].strip()
                    scenario_ws.cell(5, 7).value = row["Tester"]

                    testing_date = row["Testing Date"]
                    if not isinstance(testing_date, str) and not isinstance(testing_date, int):
                        testing_date = testing_date.strftime('%m/%d/%Y')
                    else:
                        testing_date = "30/06/2023"

                    scenario_ws.cell(5, 8).value = testing_date

                    scenario_ws.cell(5, 9).value = row["Reviewer"]

                    # handle excel date format
                    review_date = row["Review Date"]
                    if not isinstance(review_date, str) and not isinstance(review_date, int):
                        review_date = review_date.strftime('%m/%d/%Y')
                    else:
                        review_date = "07/07/2023"

                    scenario_ws.cell(5, 10).value = review_date

                    #   1.process api result screenshot
                    print(f'      pin api result screenshot')
                    scenario_ws.cell(10, 1).value = "API Test Result"

                    if api_id == "API-007":
                        api_result_image_file_path = f'{folder_input_api_result}/{api_id}/{row["Scenario"]}.png'
                    else:
                        api_result_image_file_path = convert_file_txt_to_image(folder_input_api_result, api_id,
                                                                               row["Scenario"])
                    add_image_to_excel(api_result_image_file_path, scenario_ws, "A12")

                    #   2.process sql statement screenshot
                    print(f'      pin sql statement screenshot')
                    scenario_ws.cell(10, 6).value = "SQL Statements"

                    sql_statement_image_file_path = convert_file_txt_to_image(folder_input_sql_statement, api_id,
                                                                              row["Scenario"])
                    add_image_to_excel(sql_statement_image_file_path, scenario_ws, "F12")

                    #   3.process db schema screenshot
                    print(f'      pin db schema screenshot')
                    scenario_ws.cell(90, 1).value = "DB Schema"

                    file_names_with_extension = os.listdir(folder_input_db_schema)
                    file_names_without_extension = [os.path.splitext(file_name)[0] for file_name in
                                                    file_names_with_extension]

                    with open(f'{folder_input_sql_statement}/{api_id}/{row["Scenario"]}.txt') as file:
                        txt_content = file.read()

                    txt_content = txt_content.upper()

                    # sql statement match db table
                    matching_table_name = [file_name for file_name in file_names_without_extension if
                                           file_name in txt_content]

                    start_value = 92
                    for table_name in matching_table_name:
                        db_schema_file = f'{folder_input_db_schema}/{table_name}.png'
                        cell_reference = f"A{start_value}"
                        add_image_to_excel(db_schema_file, scenario_ws, cell_reference)

                        start_value += 70

                    #   4.process jaeger screenshot
                    print(f'      pin jaeger screenshot')
                    scenario_ws.cell(90, 12).value = "Jaeger Result"

                    jaeger_image_file_path = f'{folder_input_jaeger}/{api_id}/{row["Scenario"]}*'

                    matching_files = glob.glob(jaeger_image_file_path)

                    if matching_files:
                        for file_path in matching_files:
                            add_image_to_excel(file_path, scenario_ws, "L92")

                    else:
                        print(f'no matching jaeger screenshot - {api_id}/{row["Scenario"]}')

        wb.remove(wb['Scenario Template'])
        wb.save(output_file)
        print(f"////////////  finish  {name} ////////////\n\n")


if __name__ == '__main__':
    start()
