import json

import openpyxl

from utils.PathManager import load_path_manager as lpm

file_pm_aa = lpm.input.Postman("postman_collection.json")
file_pm_batch = lpm.input.Postman("postman_collection.json")
file_pm_common = lpm.input.Postman("postman_collection.json")
file_master_data = lpm.input("API Test Master Data.xlsx")

OCP_INTERNAL_HOST = ""
DEV_HOST = ""

cv_ROUTE = "/cas/aa/v1"
ADMIN_ROUTE = "/cas/bb/v1/system/admin"
BATCH_ROUTE = "/cas/bb/batch/v1"
COMMON_ROUTE = "/cas/bb/v1"

cv_HOST = OCP_INTERNAL_HOST + cv_ROUTE
ADMIN_HOST = OCP_INTERNAL_HOST + ADMIN_ROUTE
BATCH_HOST = OCP_INTERNAL_HOST + BATCH_ROUTE
COMMON_HOST = DEV_HOST + COMMON_ROUTE


def load_postman_file_data(postman_file):
    with open(postman_file, 'r', encoding='utf-8-sig') as file:  # Note encoding='utf-8-sig'
        postman_data = json.load(file)

    return postman_data


def find_requests_responses(data):
    api_collection = []
    if isinstance(data, dict):
        if "request" in data and "response" in data:
            api_collection.append(data)

        else:
            for key, value in data.items():
                api_collection.extend(find_requests_responses(value))
    elif isinstance(data, list):
        for item in data:
            api_collection.extend(find_requests_responses(item))

    return api_collection


def compose_collection_mapping(collection_data, api_data):
    api_count = 0
    for collection_item in collection_data:
        method_value = collection_item['request']['method']
        url_raw_value = collection_item['request']['url']['raw']
        # print(f"Method: {method_value}, URL Raw: {url_raw_value}")
        match_flag = False
        for api_item in api_data:
            if api_item['HTTP Method'] == method_value:

                clear_url_raw_value = url_raw_value.replace("{{cv-host}}", "") \
                    .replace("{{admin-host}}", "") \
                    .replace("{{batch-host}}", "") \
                    .replace("{{baseUrl}}", "") \
                    .replace("/internal", "") \
                    .replace(":", "")

                clear_api_end_point = api_item['API End Point'].replace("{", "") \
                    .replace("}", "") \
                    .replace("/cas/aa/v1", "") \
                    .replace("/cas/aa/batch/v1/internal", "") \
                    .replace("/cas/cv/v1/internal", "")

                if clear_api_end_point == clear_url_raw_value:
                    api_count += 1

                    collection_item['API ID'] = api_item['API ID']

                    match_flag = True

        if match_flag is False:
            print(f"!!! Postman -> Method: {method_value}, URL Raw: {url_raw_value} \n"
                  f"not matching Master data Excel's MASTER Sheet <API End Point>\n")

    # print(f"total api count {api_count}")
    return collection_data


def extract_data_from_master_data_excel():
    wb = openpyxl.load_workbook(file_master_data)
    ws = wb["MASTER"]

    header_row = ws[1]
    header = [cell.value for cell in header_row]

    data_list = []

    current_module_section = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        data_map = {}
        for col_idx, value in enumerate(row, start=1):

            col_name = header[col_idx - 1]

            # assort module section
            if col_idx == 1:
                if value is not None:
                    current_module_section = value
                else:
                    value = current_module_section

            data_map[col_name] = value

        data_list.append(data_map)

    wb.close()

    return data_list


def process_item(item):
    path_parameter = None
    request_body = None

    # for common user api remove /internal route
    url = item['originalRequest']['url']['raw']

    if "{{baseUrl}}" in url:
        url = url.replace("/internal", "")

    api_end_point = url.replace("{{cv-host}}", cv_ROUTE) \
        .replace("{{admin-host}}", ADMIN_ROUTE) \
        .replace("{{batch-host}}", BATCH_ROUTE) \
        .replace("{{baseUrl}}", COMMON_ROUTE)

    # Note host domain
    request_api_url = url.replace("{{cv-host}}", cv_HOST) \
        .replace("{{admin-host}}", ADMIN_HOST) \
        .replace("{{batch-host}}", BATCH_HOST) \
        .replace("{{baseUrl}}", COMMON_HOST)

    if 'variable' in item['originalRequest']['url']:
        path_variable = item['originalRequest']['url']['variable']
        formatted_variables = ['='.join([pv['key'], pv['value']]) for pv in path_variable]
        path_parameter = ','.join(formatted_variables)

        # compose variable in url
        for pv in path_variable:
            # print(pv['key'])
            api_end_point = api_end_point.replace(":" + pv['key'], "{" + pv['key'] + "}")
            request_api_url = request_api_url.replace(":" + pv['key'], "{" + pv['key'] + "}")

    if 'body' in item['originalRequest']:
        request_body = item['originalRequest']['body']['raw']

    # print(request_api_url)
    return path_parameter, request_body, api_end_point, request_api_url


def fill_each_api_scenario_to_master_data(collection_data):
    wb = openpyxl.load_workbook(file_master_data)

    for api_item in collection_data:

        if "API ID" not in api_item:
            print(
                f"!!! not matching api id {api_item['name']}, check -> function compose_collection_mapping()")
            continue

        ws = wb[api_item["API ID"]]

        header_map = {col_idx: cell.value for col_idx, cell in enumerate(ws[1], 0)}

        scenario_name_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            scenario = row[9]  # Scenario column
            scenario_name_list.append(scenario)

        scenario_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            scenario_list.append(row)

        scenario_list_result = []

        print(f"processing <excel sheet {api_item['API ID']}> / <postman request {api_item['name']}>")
        for item in api_item["response"]:

            item_name = item["name"]

            if item_name in scenario_name_list:
                #  exist same scenario , update it
                print(f"        example: ✔ exist same scenario {item_name} ")

                for row in scenario_list:
                    if row[9] == item_name:
                        # 'Path Parameter(s)'   15
                        # 'Request Body'        16
                        # 'Response Body (RAW)' 22

                        path_parameter, request_body, api_end_point, request_api_url = process_item(item)

                        row = list(row)
                        # can add more field to excel in here
                        row[4] = api_end_point
                        row[13] = request_api_url
                        row[15] = path_parameter
                        row[16] = request_body

                        scenario_list_result.append(row)

            else:
                #  new scenario , add it
                print(f"        example: new scenario {item_name}, add to excel")

                path_parameter, request_body, api_end_point, request_api_url = process_item(item)

                new_scenario_list = [None] * 35
                # can add more field to excel in here
                row[4] = api_end_point
                row[13] = request_api_url
                new_scenario_list[9] = item_name
                new_scenario_list[15] = path_parameter
                new_scenario_list[16] = request_body
                new_scenario = tuple(new_scenario_list)

                scenario_list_result.append(new_scenario)

        for row_idx, row_data in enumerate(scenario_list_result, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        print(f"finish ------------------------------------------------------------------------")

    wb.save(file_master_data)
    wb.close()


def start():
    # 1.get postman export file data
    cv_data = load_postman_file_data(file_pm_aa)
    batch_data = load_postman_file_data(file_pm_batch)
    common_data = load_postman_file_data(file_pm_common)

    # 2.get postman file request and response example data
    cv_collection = find_requests_responses(cv_data)
    batch_collection = find_requests_responses(batch_data)
    common_collection = find_requests_responses(common_data)

    # 3.get api id and api endpoint mapping
    api_data = extract_data_from_master_data_excel()

    # 4.compose api id
    # cv_collection_mapping = compose_collection_mapping(cv_collection, api_data)
    # batch_collection_mapping = compose_collection_mapping(batch_collection, api_data)
    common_collection_mapping = compose_collection_mapping(common_collection, api_data)

    # 5.fill postman example / api scenario to master data excel
    # fill_each_api_scenario_to_master_data(cv_collection_mapping)
    # fill_each_api_scenario_to_master_data(batch_collection_mapping)
    fill_each_api_scenario_to_master_data(common_collection_mapping)


if __name__ == '__main__':
    start()
